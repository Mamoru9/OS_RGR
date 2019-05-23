import xlrd, openpyxl
from openpyxl.styles import PatternFill
from array import *


wb = openpyxl.load_workbook(filename = 'text.xlsx')
sheet = wb['Лист1']

data = array('i',[])
for i in range(500):
    data.insert(i, 0)

temp = 0
flag = True

while (True):
    print("Чтобы выйти нажмите ctrl+c") 
    writer = int(input("Введите кол-во квантов записи: "))
    if flag:
        reader_start = int(input("Введите квант с которого начинает читать читатель: "))
        flag = False
    reader = int(input("Введите кол-во квантов чтения:"))
    while writer!=0:
        sheet.cell(row=2, column=temp+2).fill=PatternFill(start_color='008000', end_color='008000', fill_type='solid')
        if flag != True:
            sheet.cell(row=3, column=temp+2).fill=PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        writer-=1
        data[temp]=1
        temp+=1
    tmp_bool = True
    while reader!=0:
        if tmp_bool and flag:
            while True:
                if data[reader_start-1]==1:
                    sheet.cell(row=3, column=reader_start+1).fill=PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    reader_start+=1
                else:
                    tmp_bool = False
                    break

        sheet.cell(row=3, column=temp+2).fill=PatternFill(start_color='008000', end_color='008000', fill_type='solid')  
        sheet.cell(row=2, column=temp+2).fill=PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        reader-=1
        data[temp]=1
        temp+=1
    wb.save('text.xlsx')

 